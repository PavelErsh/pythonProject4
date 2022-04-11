import main
import openpyxl
import datetime
from openpyxl.styles import Font, Fill


def crate_book():
    book = openpyxl.Workbook()
    now = datetime.datetime.now()

    book.create_sheet('Футбол', 0)
    sheet = book.active


    id_list, t1_list,t2_list, name_list = main.get_data()

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)


    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'футбол' in name_list[i] or 'Футбол' in name_list[i]  and not 'FIFA' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter+=1
            id+=1


    book.create_sheet('Хоккей', 1)
    book
    sheet = book['Хоккей']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'хоккей' in name_list[i] or 'Хоккей' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1

    book.create_sheet('Гандбол', 2)
    book
    sheet = book['Гандбол']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'Гандбол' in name_list[i] or 'гандбол' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1
    book.create_sheet('Баскетбол', 3)
    book
    sheet = book['Баскетбол']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'Баскетбол' in name_list[i] or 'баскетбол' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1

    book.create_sheet('Волейбол', 4)
    book
    sheet = book['Волейбол']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'Волейбол' in name_list[i] or 'Волейбол' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1

    book.create_sheet('Большой Теннис', 5)
    book
    sheet = book['Большой Теннис']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'Теннис' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1

    book.create_sheet('Настольный Теннис', 5)
    book
    sheet = book['Настольный Теннис']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'Настольный теннис' in name_list[i] or 'настольный теннис' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1

    book.create_sheet('Киберспорт FIFA футбол', 6)
    book
    sheet = book['Киберспорт FIFA футбол']

    sheet['A1'] = 'НОМЕР ИГРЫ'
    sheet['B1'] = 'ID ИГРЫ'
    sheet['C1'] = 'ТЕКУЩЕЕ ВРЕМЯ'
    sheet['D1'] = 'ВРЕМЯ ИГРЫ'
    sheet['E1'] = 'ЛИГА'
    sheet['F1'] = 'КОМАНДА 1'
    sheet['G1'] = 'КОМАНДА 2'

    work_sheet_a1 = sheet['A1']
    work_sheet_a1.font = Font(size=15, bold=True)

    work_sheet_b1 = sheet['B1']
    work_sheet_b1.font = Font(size=15, bold=True)

    work_sheet_c1 = sheet['C1']
    work_sheet_c1.font = Font(size=15, bold=True)

    work_sheet_d1 = sheet['D1']
    work_sheet_d1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['E1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_e1 = sheet['F1']
    work_sheet_e1.font = Font(size=15, bold=True)

    work_sheet_g1 = sheet['G1']
    work_sheet_g1.font = Font(size=15, bold=True)

    work_sheet_h1 = sheet['H1']
    work_sheet_h1.font = Font(size=15, bold=True)

    counter = 2
    id = 0
    for i in range(len(name_list)):
        if 'FIFA' in name_list[i]:
            sheet[f'A{counter}'] = id
            sheet[f'B{counter}'] = id_list[i]
            sheet[f'C{counter}'] = str(now.strftime("%d-%m-%Y %H:%M"))

            sheet[f'E{counter}'] = name_list[i]
            sheet[f'F{counter}'] = t1_list[i]
            sheet[f'G{counter}'] = t2_list[i]
            counter += 1
            id += 1


    del book['Sheet']

    book.save('my_book.xlsx')
    book.close()

crate_book()

